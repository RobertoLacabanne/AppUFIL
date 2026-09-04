"""
El membrete de la hoja impresa, y el cuño de demostración.

Dos cosas que sólo importan cuando el sistema deja de estar en la pantalla:

  · Una hoja que sale de acá termina agregada a un legajo. Tiene que decir de dónde
    salió sin depender de que alguien se acuerde de escribirlo: el organismo, la
    unidad, qué legajo y cuándo se emitió.
  · Y si esa hoja trae contratos inventados, tiene que decirlo de una manera que no se
    pueda perder. El aviso era una franja horizontal que cruzaba TODAS las pantallas,
    siempre; a los dos días se aprende a no verla, que es lo peor que le puede pasar a
    un aviso de seguridad. Ahora es un cuño girado, fijo abajo a la derecha, con el
    doble filete que el sistema ya usa.

Un cuño de seguridad tiene tres condiciones y las tres se verifican acá: no se puede
cerrar, no se puede tapar, y en impresión sale sí o sí.
"""
from __future__ import annotations

import re
import sys
import unittest
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

CSS = (RAIZ / "ufil/web/estilo.css").read_text(encoding="utf-8")
HTML = (RAIZ / "ufil/web/index.html").read_text(encoding="utf-8")
APP = (RAIZ / "ufil/web/app.js").read_text(encoding="utf-8")


def _bloque_impresion() -> str:
    """El `@media print` grande, con las llaves balanceadas."""
    marca = "@media print{"
    i, mejor = 0, ""
    while (i := CSS.find(marca, i)) != -1:
        j, hondo = i + len(marca), 1
        while j < len(CSS) and hondo:
            hondo += (CSS[j] == "{") - (CSS[j] == "}")
            j += 1
        trozo = CSS[i + len(marca):j - 1]
        if len(trozo) > len(mejor):
            mejor = trozo
        i = j
    return mejor


class LaHojaImpresaDiceDeDondeSalio(unittest.TestCase):

    def test_el_membrete_existe_y_solo_en_papel(self):
        self.assertIn('id="membrete"', HTML, "se perdió el membrete de impresión")
        self.assertRegex(
            CSS, r"#membrete\{display:none\}",
            "el membrete se ve en pantalla: la identidad ya está en la barra lateral "
            "y repetirla arriba de cada folio es decir dos veces lo mismo")
        self.assertIn("#membrete{display:flex", _bloque_impresion(),
                      "el membrete dejó de aparecer al imprimir")

    def test_dice_las_cuatro_cosas(self):
        for id_, que in (("membrete-organismo", "el organismo"),
                         ("membrete-unidad", "la unidad"),
                         ("membrete-legajo", "qué legajo"),
                         ("membrete-fecha", "cuándo se emitió")):
            self.assertIn(f'id="{id_}"', HTML,
                          f"el membrete dejó de decir {que}")

    def test_el_nombre_sale_de_identidad_y_no_escrito_a_mano(self):
        self.assertIn("$('#membrete-organismo')", APP,
                      "el membrete dejó de tomar el nombre del endpoint de identidad: "
                      "cambiar de unidad dejaría hojas impresas con el nombre viejo")

    def test_la_hora_se_sella_al_imprimir_y_no_al_cargar(self):
        """
        Una pestaña abierta desde la mañana imprimiría la hora de la mañana, y esa
        hoja se agrega a un legajo.
        """
        self.assertIn("addEventListener('beforeprint'", APP,
                      "la fecha de emisión volvió a fijarse al cargar la página")


class ElCunoDeDemostracionNoSePuedePerder(unittest.TestCase):

    def _regla(self, donde=CSS):
        m = re.search(r"#aviso-demo\{([^{}]*)\}", donde)
        self.assertIsNotNone(m, "se perdió el cuño de demostración")
        return m.group(1).replace(" ", "")

    def test_es_un_cuno_y_no_una_franja(self):
        cuerpo = self._regla()
        self.assertIn("position:fixed", cuerpo,
                      "volvió a ser una franja en el flujo, que se come un renglón de "
                      "todas las pantallas para siempre")
        self.assertRegex(cuerpo, r"transform:rotate\(-?\d",
                         "el cuño dejó de estar girado y se lee como un cartel más")

    def test_no_estorba_el_trabajo_pero_no_se_puede_cerrar(self):
        cuerpo = self._regla()
        self.assertIn("pointer-events:none", cuerpo,
                      "el cuño intercepta clics: puesto encima del trabajo, tapa lo "
                      "que hay debajo")
        # No se puede cerrar porque no hay con qué: ningún botón en el marcado.
        i = HTML.index('id="aviso-demo"')
        self.assertNotIn("<button", HTML[i:HTML.index("</div>", i)],
                         "le pusieron un botón de cerrar a un aviso de seguridad")

    def test_en_papel_sale_si_o_si(self):
        impresion = _bloque_impresion()
        self.assertIn("#aviso-demo{position:static", impresion.replace(" ", ""),
                      "el cuño no está resuelto para el papel: una hoja con contratos "
                      "inventados podría salir sin decirlo")
        self.assertNotRegex(
            impresion, r"#aviso-demo\{[^{}]*display:none",
            "el cuño se apaga al imprimir, que es justo cuando más hace falta")

    def test_el_texto_largo_no_esta_adentro_del_cuno(self):
        """Un cuño es breve o no es un cuño: el párrafo entero lo volvía un cartel."""
        i = HTML.index('id="aviso-demo"')
        marcado = HTML[i:HTML.index("</div>", i)]
        self.assertIn("title=", marcado,
                      "la explicación completa dejó de estar al alcance")
        cuerpo = re.sub(r"<[^>]*>", " ", marcado.split(">", 1)[1])
        self.assertLess(len(" ".join(cuerpo.split())), 90,
                        "el cuño volvió a llevar el párrafo entero adentro")


class UnaTablaImpresaNoDesperdiciaLaPrimeraHoja(unittest.TestCase):
    """
    Medido imprimiendo la pantalla de contratos —51 filas— a A4: salían **cinco
    hojas y la primera en blanco**, con el membrete y nada más, y la tabla empezaba en
    la segunda.

    La causa: «ningún bloque se parte». Con una tabla más alta que una hoja el
    navegador hace lo único que puede —empezarla en la página siguiente y partirla
    igual—, así que la regla no evitaba nada y costaba una hoja en cada impresión. Y
    la primera hoja de algo que se agrega a un legajo, en blanco.

    Lo que no se parte es lo CHICO y el RENGLÓN de una tabla. Lo grande fluye, que es
    lo que hace cualquier tabla impresa desde que existen.

    Después del cambio: tres hojas, contenido desde el primer renglón.
    """

    def setUp(self):
        self.papel = _bloque_impresion()

    def _regla(self, selector):
        m = re.search(re.escape(selector) + r"\{([^{}]*)\}", self.papel)
        self.assertIsNotNone(m, f"se perdió la regla «{selector}» de impresión")
        return m.group(1).replace(" ", "").replace("\n", "")

    def test_un_bloque_grande_puede_partirse(self):
        self.assertIn("break-inside:auto", self._regla(".bloque"),
                      "volvió «ningún bloque se parte»: con una tabla más alta que "
                      "una hoja eso no evita nada y desperdicia la primera")
        self.assertNotIn("page-break-inside:avoid", self._regla(".bloque"))

    def test_pero_un_renglon_no_se_parte_por_la_mitad(self):
        self.assertIn("break-inside:avoid", self._regla("tbody tr"),
                      "un renglón partido entre dos hojas deja media fila de datos "
                      "en cada una")

    def test_el_encabezado_se_repite_en_cada_hoja(self):
        """Una tabla partida sin encabezado obliga a volver a la hoja anterior para
        saber qué columna es cuál."""
        self.assertIn("display:table-header-group", self._regla("thead"))

    def test_lo_chico_sigue_entero(self):
        c = self._regla(".cifras, .carriles, .cronologia, .interp")
        self.assertIn("break-inside:avoid", c,
                      "una baldosa de cifras o una tarjeta de interpretación partida "
                      "entre dos hojas no se lee")

    def test_un_titulo_no_se_queda_solo_al_pie(self):
        self.assertIn("break-after:avoid", self._regla("h2, h3"))

    def test_no_se_fijan_los_margenes_de_la_hoja(self):
        """
        Se probó fijarlos —14 mm, 18 a la izquierda para el perforado— y la tabla de
        contratos, que ya venía justa, quedaba CORTADA a la derecha: se perdía la
        columna de confianza. Un margen más lindo no vale una columna menos.
        """
        self.assertNotIn("@page{", CSS.replace(" ", ""),
                         "volvieron los márgenes fijos: la tabla de contratos se "
                         "corta a la derecha")


class LaMarcaEnLoQueSeExporta(unittest.TestCase):
    """
    El `.rtf` y el `.xlsx` que salen del sistema son lo que después se pega en un
    escrito, o se adjunta. Ahí el encabezado del organismo no es adorno: es lo que
    hace que la hoja se sostenga sola cuando la mira alguien de afuera.

    Tres cosas se verifican acá:

      · que esté, y arriba del título, que es donde va un membrete;
      · que las líneas salgan de `ufil/identidad.py` y no escritas a mano —escritas
        a mano, cambiar de fiscal deja el nombre viejo adentro de un documento ya
        entregado, que es exactamente lo que no se puede permitir—;
      · que se pueda sacar, porque un borrador interno no tiene por qué salir con
        la marca oficial encima.
    """

    def setUp(self):
        import tempfile
        from ufil import db
        from ufil.db import ahora
        self.tmp = tempfile.TemporaryDirectory()
        self.cx = db.abrir(Path(self.tmp.name) / "t.sqlite")
        self.cx.execute("""INSERT INTO archivo (sha256,ruta_original,nombre,bytes,ingerido_en)
                           VALUES ('aa','/x/a.pdf','a.pdf',1,?)""", (ahora(),))
        self.cx.execute("INSERT INTO documento (sha256,tipo,perfil) VALUES ('aa','contrato_obra','p')")
        self.cx.commit()

    def tearDown(self):
        self.cx.close(); self.tmp.cleanup()

    def _rtf(self, **kw) -> str:
        from ufil import capa7_export as c7
        destino = Path(self.tmp.name) / "i.rtf"
        return Path(c7.a_rtf(self.cx, destino, **kw)).read_text(encoding="cp1252")

    def _portada(self, **kw) -> str:
        import openpyxl
        from ufil import capa4_analisis as c4, capa7_export as c7
        destino = Path(self.tmp.name) / "a.xlsx"
        c7.a_xlsx(self.cx, destino, [c["id"] for c in c4.catalogo()], **kw)
        hoja = openpyxl.load_workbook(destino)["procedencia"]
        return " ".join(str(c) for f in hoja.iter_rows(values_only=True)
                        for c in f if c is not None)

    def test_el_informe_sale_encabezado_y_arriba_del_titulo(self):
        from ufil import capa7_export as c7
        from ufil import identidad as ident
        texto = self._rtf()
        primera = c7._rtf(ident.encabezado_export()[0])
        self.assertIn(primera, texto, "el .rtf sale sin el encabezado del organismo")
        # Un membrete abajo del título no es un membrete.
        self.assertLess(texto.index(primera), texto.index("INFORME DE AN"),
                        "el encabezado tiene que ir arriba del título")

    def test_el_informe_lleva_a_los_fiscales(self):
        from ufil import capa7_export as c7
        from ufil import identidad as ident
        firma = ident.firma()
        self.assertTrue(firma, "la identidad de la casa tiene que traer fiscales")
        self.assertIn(c7._rtf(firma), self._rtf())

    def test_los_nombres_salen_de_identidad_y_no_escritos_a_mano(self):
        """
        Se cambia el organismo por el entorno y tiene que cambiar el papel. Si esto
        falla es porque alguien escribió el nombre adentro de capa7_export.py, y ese
        nombre va a sobrevivir al cambio de fiscal.
        """
        import os
        from ufil import capa7_export as c7
        previo = os.environ.get("UFIL_ORGANISMO")
        os.environ["UFIL_ORGANISMO"] = "Fiscalia Inventada de Prueba"
        try:
            texto = self._rtf()
            portada = self._portada()
        finally:
            if previo is None:
                os.environ.pop("UFIL_ORGANISMO", None)
            else:
                os.environ["UFIL_ORGANISMO"] = previo
        self.assertIn(c7._rtf("Fiscalia Inventada de Prueba"), texto,
                      "el .rtf no lee el organismo de ufil/identidad.py")
        self.assertIn("Fiscalia Inventada de Prueba", portada,
                      "la portada del .xlsx no lee el organismo de ufil/identidad.py")

    def test_se_puede_sacar_para_un_borrador(self):
        from ufil import capa7_export as c7
        from ufil import identidad as ident
        primera = c7._rtf(ident.encabezado_export()[0])
        sin = self._rtf(membrete=False)
        self.assertNotIn(primera, sin, "pidieron sin membrete y salió con membrete")
        # Y sin membrete sigue siendo el informe completo, no un archivo mutilado.
        self.assertIn("INFORME DE AN", sin)
        self.assertIn("Generado el", sin)
        portada = self._portada(membrete=False)
        self.assertNotIn(ident.encabezado_export()[0], portada)
        self.assertIn("Generado", portada, "la portada perdió la procedencia")

    def test_la_portada_del_excel_lo_lleva(self):
        from ufil import identidad as ident
        portada = self._portada()
        for linea in ident.encabezado_export():
            self.assertIn(linea, portada)

    def test_el_escudo_va_en_la_portada_si_lo_pusieron(self):
        """
        Si hay `assets/marca/logo.png`, va arriba de la portada. Si no hay, la
        planilla sale igual: que falte el emblema es un problema menor, que no
        salga la planilla es un problema mayor.
        """
        import tempfile
        import openpyxl
        from PIL import Image
        from ufil import capa4_analisis as c4, capa7_export as c7, config
        marca = Path(tempfile.mkdtemp())
        Image.new("RGB", (200, 100), "white").save(marca / "logo.png")
        previo = config.MARCA
        destino = Path(self.tmp.name) / "con-escudo.xlsx"
        try:
            config.MARCA = marca
            c7.a_xlsx(self.cx, destino, [c["id"] for c in c4.catalogo()])
        finally:
            config.MARCA = previo
        hoja = openpyxl.load_workbook(destino)["procedencia"]
        self.assertEqual(len(hoja._images), 1, "el escudo no llegó a la portada")
        # Al releer el .xlsx, `img.width` vuelve a salir del PNG y no del dibujo; el
        # tamaño con el que se ve está en el ancla, en EMU.
        from openpyxl.utils.units import EMU_to_pixels
        ancla = hoja._images[0].anchor
        ext = ancla.ext
        # A la derecha del encabezado, que se desborda de la columna A.
        self.assertGreaterEqual(ancla._from.col, 5,
                                "el escudo le queda encima al nombre del organismo")
        # Alto fijo y ancho proporcional: un escudo estirado es peor que ninguno.
        self.assertEqual(EMU_to_pixels(ext.cy), 64)
        self.assertEqual(EMU_to_pixels(ext.cx), 128)

    def test_sin_escudo_la_planilla_sale_igual(self):
        import tempfile
        import openpyxl
        from ufil import capa4_analisis as c4, capa7_export as c7, config
        previo = config.MARCA
        destino = Path(self.tmp.name) / "sin-escudo.xlsx"
        try:
            config.MARCA = Path(tempfile.mkdtemp())
            c7.a_xlsx(self.cx, destino, [c["id"] for c in c4.catalogo()])
        finally:
            config.MARCA = previo
        self.assertEqual(len(openpyxl.load_workbook(destino)["procedencia"]._images), 0)

    def test_la_pantalla_ofrece_sacarlo_y_el_servidor_lo_entiende(self):
        """La opción no sirve si no llega hasta el archivo que se baja."""
        servidor = (RAIZ / "ufil/servidor.py").read_text(encoding="utf-8")
        self.assertIn('id="con-membrete"', APP, "la pantalla no ofrece la opción")
        self.assertIn("&membrete=no", APP, "la opción no viaja en el enlace")
        self.assertIn('q.get("membrete"', servidor, "el servidor ignora la opción")
        # Y no sale impresa: los botones que acompaña no se imprimen, así que en papel
        # quedaría una casilla tildada sola en medio de una hoja que va a un legajo.
        self.assertIn("opcion-suelta", _bloque_impresion(),
                      "la casilla se imprime, y en papel no manda nada")
        # Las DOS exportaciones tienen que recibirla. Pasándosela sólo al .rtf, quien
        # destildaba la casilla bajaba igual una planilla con la marca oficial encima.
        for llamada in ("c7.a_rtf(", "c7.a_xlsx("):
            i = servidor.index(llamada) + len(llamada) - 1
            j, hondo = i + 1, 1
            while j < len(servidor) and hondo:
                hondo += (servidor[j] == "(") - (servidor[j] == ")")
                j += 1
            self.assertIn("membrete", servidor[i:j],
                          f"{llamada}…) del servidor no recibe la opción")
