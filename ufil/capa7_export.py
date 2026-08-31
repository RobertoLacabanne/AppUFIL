"""
Capa 7 — Exportación.

Tablas a .xlsx y cuerpo a .rtf con interlineado 1,5, justificado y cuerpo 11.

Cada afirmación exportada arrastra su archivo y su foja. Un informe que dice
"contrató en las dos cámaras" sin decir en qué papel lo leyó no sirve para
trabajar: obliga a rehacer la búsqueda a mano.
"""
from __future__ import annotations

import sqlite3
from datetime import date

# Nombres legibles de los campos. Un informe que va a un legajo no dice `fecha_inicio`.
ROTULO_CAMPO = {
    "nombre": "contratado", "documento": "documento", "cargo": "cargo",
    "fecha_inicio": "fecha de inicio", "fecha_fin": "fecha de finalización",
    "fecha_contrato": "fecha del contrato", "monto": "monto mensual",
    "monto_total": "monto total", "monto_total_letras": "monto total en letras",
    "plazo_meses": "plazo en meses", "comprobante": "número de comprobante",
}

# En una factura los mismos campos dicen otra cosa: `nombre` es quien la emitió y
# `fecha_inicio` es la fecha de emisión. Ver docs/11-contrato-o-factura.md.
ROTULO_CAMPO_POR_FAMILIA = {
    "comprobante": {"nombre": "emisor", "documento": "CUIT del emisor",
                    "fecha_inicio": "fecha de emisión", "monto": "importe"},
    "acto": {"nombre": "título o referencia", "documento": "número o identificador",
             "fecha_inicio": "fecha"},
}
FAMILIA_ROTULO = {"contrato": "Contrato", "comprobante": "Comprobante de pago",
                  "acto": "Acto administrativo"}


def _rotulo(campo: str, familia: str | None) -> str:
    return (ROTULO_CAMPO_POR_FAMILIA.get(familia, {}).get(campo)
            or ROTULO_CAMPO.get(campo, campo))


def _periodo(texto: str | None) -> str:
    """«2023-03-01 → 2023-08-31» pasa a «01/03/2023 → 31/08/2023»."""
    if not texto:
        return "—"
    return " → ".join(fecha(t.strip()) for t in str(texto).split("→"))
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import capa4_analisis as c4
from . import confianza as cf
from .castellano import camara, concordar, fecha, pesos, plural

TINTA = "FF1B1D21"; SELLO = "FF23477A"; FILETE = "FFCFCBC2"; PAPEL2 = "FFF1EFEA"


def _hoja(wb: Workbook, res: dict) -> None:
    ws = wb.create_sheet(res["id"][:31])
    cols = res["columnas"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        celda = ws.cell(row=1, column=c)
        celda.font = Font(bold=True, size=9, color="FFFFFFFF")
        celda.fill = PatternFill("solid", fgColor=SELLO)
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    for fila in res["filas"]:
        ws.append([fila[c] for c in cols])

    fino = Side(style="thin", color=FILETE)
    for f in ws.iter_rows(min_row=2):
        for celda in f:
            celda.font = Font(size=10)
            celda.border = Border(bottom=fino)
            if isinstance(celda.value, (int, float)):
                celda.alignment = Alignment(horizontal="right")
                if "centavos" in str(ws.cell(row=1, column=celda.column).value):
                    celda.number_format = '#,##0.00'
                    celda.value = celda.value / 100.0
    for i, c in enumerate(cols, start=1):
        ancho = max(len(str(c)), *(len(str(f[c])) for f in res["filas"])) if res["filas"] else len(c)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(ancho + 2, 10), 42)
    ws.freeze_panes = "A2"


def a_xlsx(cx: sqlite3.Connection, destino: Path, consultas: list[str]) -> Path:
    wb = Workbook(); wb.remove(wb.active)
    portada = wb.create_sheet("procedencia", 0)
    portada.append(["Sistema de análisis documental — UFIL Paraná"])
    portada["A1"].font = Font(bold=True, size=13)
    portada.append([])
    portada.append(["Generado", date.today().isoformat()])
    portada.append(["Archivos ingeridos",
                    cx.execute("SELECT COUNT(*) FROM archivo").fetchone()[0]])
    # Desagregado por tipo: «12 documentos» no dice si son doce contratos o dos
    # contratos y diez facturas, que es una diferencia que cambia lo que se lee después.
    def cuenta(sql):
        return cx.execute(sql).fetchone()[0]

    portada.append(["Documentos extraídos", cuenta("SELECT COUNT(*) FROM documento")])
    portada.append(["· contratos",
                    cuenta("SELECT COUNT(*) FROM v_documento_todo WHERE familia='contrato'")])
    portada.append(["· facturas y recibos",
                    cuenta("SELECT COUNT(*) FROM v_documento_todo WHERE familia='comprobante'")])
    portada.append(["· otros documentos",
                    cuenta("""SELECT COUNT(*) FROM v_documento_todo
                               WHERE familia IS NULL OR familia='acto'""")])
    # OJO: esto filtraba por estado='a_revisar', que dejó de existir cuando entraron los
    # ocho estados de ufil/confianza.py. La consulta seguía corriendo sin error y
    # devolvía SIEMPRE cero: la planilla que se lleva el fiscal decía «0 campos
    # pendientes de revisión» con la cola llena. Una afirmación falsa en un documento
    # que se firma. Ahora sale de la lista de estados, que es una sola y está en Python.
    portada.append(["Campos pendientes de revisión",
                    cuenta(f"SELECT COUNT(*) FROM campo WHERE estado IN ({cf.SQL_PENDIENTES})")])
    portada.append(["Campos verificados por una persona",
                    cuenta(f"SELECT COUNT(*) FROM campo WHERE estado IN ({cf.SQL_HUMANOS})")])
    # Los archivos que entraron y no produjeron ningún contrato tienen que estar en la
    # portada. Si no, quien lee la planilla la toma por el panorama completo del corpus
    # y no se entera de lo que no está adentro. Un documento que se pierde en silencio
    # es lo peor que puede hacer un sistema que existe para no perder documentos.
    afuera = cx.execute("""SELECT COUNT(*) FROM archivo a
                            WHERE NOT EXISTS (SELECT 1 FROM documento d
                                               WHERE d.sha256 = a.sha256)""").fetchone()[0]
    portada.append(["Archivos que NO dieron ningún documento", afuera])
    portada.append([])
    portada.append(["Advertencia:"])
    portada.append(["Los valores de estas planillas se leyeron automáticamente de los "
                    "documentos. Los campos"])
    portada.append(["pendientes de revisión NO están verificados por una persona. "
                    "Antes de incorporar cualquier"])
    portada.append(["dato a un legajo, verificarlo contra el original citado."])
    if afuera:
        portada.append([])
        portada.append([f"Atención: {plural(afuera, 'archivo cargado', 'archivos cargados')} "
                        f"no {concordar(afuera, 'produjo', 'produjeron')} ningún "
                        f"documento y por lo tanto"])
        portada.append([f"NO {concordar(afuera, 'está representado', 'están representados')} "
                        "en ninguna de estas planillas. En la pantalla «Quedaron"])
        portada.append(["afuera» del sistema está la lista con el motivo de cada uno."])
    portada.column_dimensions["A"].width = 34
    portada.column_dimensions["B"].width = 22

    for cid in consultas:
        _hoja(wb, c4.correr(cx, cid))
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


# ─────────────────────────────────────────────────────────────────────── RTF ──
def _rtf(s: str) -> str:
    out = []
    for ch in str(s):
        if ch == "\\": out.append(r"\\")
        elif ch == "{": out.append(r"\{")
        elif ch == "}": out.append(r"\}")
        elif ord(ch) < 128: out.append(ch)
        else: out.append(r"\u%d?" % ord(ch))
    return "".join(out)


def a_rtf(cx: sqlite3.Connection, destino: Path) -> Path:
    """Cuerpo justificado, interlineado 1,5, cuerpo 11. Cada afirmación con su cita."""
    sup = c4.correr(cx, "01_superposicion")["filas"]
    amb = c4.correr(cx, "03_ambas_camaras")["filas"]
    cob = c4.correr(cx, "05_cobertura")["filas"]
    exc = c4.correr(cx, "06_excluidos_del_cruce")["filas"]

    P = r"\pard\qj\sl360\slmult1\fi567\sa120\fs22 "
    T = r"\pard\qc\sl360\slmult1\sa200\b\fs26 "
    H = r"\pard\ql\sl360\slmult1\sa120\sb200\b\fs24 "

    p = [r"{\rtf1\ansi\ansicpg1252\deff0",
         r"{\fonttbl{\f0\froman\fcharset0 Times New Roman;}}",
         r"\margl1701\margr1134\margt1134\margb1134",
         T + _rtf("INFORME DE ANÁLISIS DOCUMENTAL") + r"\b0\par",
         P + _rtf(f"Generado el {date.today().strftime('%d/%m/%Y')} sobre "
                  f"{plural(cx.execute('SELECT COUNT(*) FROM documento').fetchone()[0], 'documento procesado', 'documentos procesados')} "
                  f"automáticamente. Cada afirmación de este informe cita el "
                  f"archivo y la foja de donde se leyó el dato, para poder verificarla "
                  f"contra el original.") + r"\par"]

    p.append(H + _rtf("1. Alcance y cobertura de la lectura") + r"\b0\par")
    for c in cob:
        # Con el tipo de documento adelante: el campo «nombre» de un contrato y el de
        # una factura salían los dos como «Campo contratado», tres párrafos iguales con
        # números distintos y sin forma de saber cuál era cuál.
        p.append(P + _rtf(
            f"{FAMILIA_ROTULO.get(c['familia'], 'Documento sin clasificar')} · campo "
            f"«{_rotulo(c['campo'], c['familia'])}»: {c['firmes']} de {c['total']} en "
            f"estado firme ({c['pct_firme_sobre_total']}%), de los cuales "
            f"{c['verificados_por_persona']} fueron verificados por una persona; "
            f"{plural(c['pendientes_baja_confianza'], 'pendiente', 'pendientes')} por "
            f"confianza baja; {c['conflictos']} en conflicto entre rutas de lectura; "
            f"{c['sin_revisar']} sin lectura. Sólo los firmes entran en las cifras de "
            f"este informe.") + r"\par")
    if exc:
        p.append(P + r"\i " + _rtf(
            f"{concordar(len(exc), 'Quedó', 'Quedaron')} fuera del cruce de superposición "
            f"{plural(len(exc), 'contrato', 'contratos')} por "
            f"faltarle{concordar(len(exc), '', 's')} algún dato firme. Se "
            f"{concordar(len(exc), 'detalla', 'detallan')} en la planilla adjunta: el "
            f"total de hallazgos no debe leerse como si el universo estuviera completo.")
            + r"\i0\par")
    # Y lo otro que falta: los archivos que ni siquiera llegaron a ser un contrato. Sin
    # esta línea, el informe se lee como si el universo fueran los documentos que sí
    # entraron, y quien lo lee no tiene forma de saber que hay papel afuera.
    afuera = cx.execute("""SELECT COUNT(*) FROM archivo a
                            WHERE NOT EXISTS (SELECT 1 FROM documento d
                                               WHERE d.sha256 = a.sha256)""").fetchone()[0]
    if afuera:
        p.append(P + r"\i " + _rtf(
            f"Además, {plural(afuera, 'archivo cargado', 'archivos cargados')} no "
            f"{concordar(afuera, 'produjo', 'produjeron')} ningún documento —porque no "
            f"se pudieron abrir, o porque ninguna de sus fojas se reconoció como un "
            f"formulario conocido— y por lo tanto no "
            f"{concordar(afuera, 'está representado', 'están representados')} en ninguna "
            f"cifra de este informe. El motivo de cada uno figura en el sistema.")
            + r"\i0\par")

    p.append(H + _rtf("2. Superposición temporal de contratos") + r"\b0\par")
    if not sup:
        p.append(P + _rtf("No se detectaron superposiciones entre los contratos legibles.") + r"\par")
    for s in sup[:60]:
        p.append(P + _rtf(
            f"{s['contratado']} (documento {s['documento'] or 'no legible'}) registra "
            f"contratos superpuestos por {plural(s['dias_solapados'], 'día', 'días')}, "
            f"de tipo {s['cruce']}: el período {_periodo(s['periodo_a'])} en la cámara de "
            f"{camara(s['camara_a'])} y el período {_periodo(s['periodo_b'])} en la de "
            f"{camara(s['camara_b'])} (cf. {s['archivo_a']} y {s['archivo_b']}).") + r"\par")

    p.append(H + _rtf("3. Contratados presentes en ambas cámaras") + r"\b0\par")
    if not amb:
        p.append(P + _rtf("No se detectaron contratados en ambas cámaras.") + r"\par")
    for a in amb[:60]:
        # OJO con lo que había acá: el `.replace(",", "@").replace(".", ",")…` que daba
        # vuelta el separador decimal se aplicaba a la FRASE ENTERA, no al número. El
        # informe salía diciendo «PEREZ ROMERO. Ana Laura» y «cámara B. entre 2023…».
        # Un documento que se firma, con los apellidos mal escritos. El formato del
        # importe vive en ufil/castellano.py y no toca el resto del texto.
        p.append(P + _rtf(
            f"{a['contratado']} (documento {a['documento'] or 'no legible'}) registra "
            f"{plural(a['contratos_camara_a'], 'contrato', 'contratos')} en la cámara de "
            f"{camara('A')} y {a['contratos_camara_b']} en la de {camara('B')}, entre "
            f"{fecha(a['desde'])} y {fecha(a['hasta'])}, por un acumulado mensual de "
            f"{pesos(a['acumulado_centavos'])}.") + r"\par")

    p.append(H + _rtf("Advertencia") + r"\b0\par")
    p.append(P + _rtf(
        "Este informe es una herramienta de trabajo interna. Los datos fueron leídos "
        "automáticamente y los campos pendientes de revisión no están verificados por una "
        "persona. Nada de lo que aquí se afirma debe incorporarse a un legajo sin cotejarlo "
        "antes contra la documentación original citada.") + r"\par")
    p.append("}")

    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text("\n".join(p), encoding="cp1252", errors="replace")
    return destino


def exportar(cx: sqlite3.Connection, destino: Path, consultas: list[str] | None = None) -> list[str]:
    destino = Path(destino); destino.mkdir(parents=True, exist_ok=True)
    consultas = consultas or [c["id"] for c in c4.catalogo()]
    hechos = [str(a_xlsx(cx, destino / "analisis.xlsx", consultas)),
              str(a_rtf(cx, destino / "informe.rtf"))]
    return hechos
